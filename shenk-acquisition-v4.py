#pip install sfepyimport tkinter as tk
import tkinter as tk 
from tkinter import filedialog, messagebox
from tkinter import ttk
import serial
from serial.tools import list_ports
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from openpyxl.drawing.image import Image 
import os


class App:
    def __init__(self, root):
        self.serial_obj = None
        self.stop = True
        self.enter = True
        self.zeroforce = 0
        self.zeropos = 0
        self.arrayp = []
        self.arrayf = []
        self.start_time = None  # Variabile per memorizzare l'orario di inizio prova
        self.connection_attempt_start_time = None  # Tempo dell'ultimo tentativo di connessione

        # Get the screen dimensions for adaptive sizing
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = int(screen_width * 0.9)
        window_height = int(screen_height * 0.8)

        # GUI setup with adaptive size
        root.title("Shenk App")
        root.geometry(f"{window_width}x{window_height}")

        # Define layout frames for better control
        self.main_frame = ttk.Frame(root)
        self.main_frame.grid(sticky=tk.NSEW)

        root.grid_columnconfigure(0, weight=1)
        root.grid_rowconfigure(0, weight=1)

        # Usage instructions
        self.instructions_label = ttk.Label(self.main_frame, text="1) Collegare USB \n2) Selezionare da menù a tendina la 'PORTA COM' che compare (es: COM3)\n3) Avviare la prova", font=("Arial", 12))
        self.instructions_label.grid(row=2, column=6, columnspan=5, padx=5, pady=5, sticky=tk.W)

        # Serial connection status label
        self.conn_status_label = ttk.Label(self.main_frame, text="Ricerca connessione...", font=("Arial", 12))
        self.conn_status_label.grid(row=1, column=6, columnspan=5, padx=5, pady=5, sticky=tk.W)

        # Dropdown for selecting COM ports
        self.com_port_label = ttk.Label(self.main_frame, text="Seleziona Porta COM:", font=("Arial", int(13 * 1.3)))
        self.com_port_label.grid(row=0, column=6, padx=5, pady=5, sticky=tk.E)

        self.com_ports = self.get_com_ports()
        self.com_port_var = tk.StringVar(value=self.com_ports[0] if self.com_ports else "")
        self.com_port_dropdown = ttk.Combobox(self.main_frame, textvariable=self.com_port_var, values=self.com_ports)
        self.com_port_dropdown.grid(row=0, column=7, padx=5, pady=5, sticky=tk.W)

            # Associa la selezione della porta COM all'evento di connessione
        self.com_port_dropdown.bind("<<ComboboxSelected>>", self.connect_serial)

        # Inizializza le variabili

        self.abs_pos_var = tk.DoubleVar(value=0.0)  # Variabile per "Posizione Abs."
        self.pos_var = tk.DoubleVar(value=0.0)      # Variabile per "Posizione"
        self.carico_var = tk.DoubleVar(value=0.0)   # Variabile per "Carico"

        # Create position and carico entry fields
        self.create_numeric_entry("Posizione Abs. [mm]:", 0, 1.3, self.abs_pos_var, editable=False)
        self.create_numeric_entry("Posizione [mm]:", 1, 1.3, self.pos_var, editable=False)
        self.create_numeric_entry("Carico [Kg]:", 2, 1.3, self.carico_var, editable=False)

        """
        self.create_numeric_entry('Posizione Abs. [mm]:', 0, 1.3, editable=False)
        self.create_numeric_entry('Posizione [mm]:', 1, 1.3, editable=False)
        self.create_numeric_entry('Carico [Kg]:', 2, 1.3, editable=False)
        """

        # Speed of test field
        self.speed_label = ttk.Label(self.main_frame, text="Velocità della prova [mm/min]:", font=("Arial", int(13 * 1.3)))
        self.speed_label.grid(row=3, column=0, padx=5, pady=5, sticky=tk.E)
        self.speed_var = tk.DoubleVar()
        self.speed_entry = ttk.Entry(self.main_frame, textvariable=self.speed_var, font=("Arial", int(16 * 1.3)))
        self.speed_entry.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)

        # Control buttons
        self.start_button = ttk.Button(self.main_frame, text="Test Start", command=self.start_test, style="Green.TButton")
        self.start_button.grid(row=2, column=4, padx=10, pady=5, sticky=tk.EW)

        self.stop_button = ttk.Button(self.main_frame, text="Test Stop", command=self.stop_test, style="Red.TButton")
        self.stop_button.grid(row=3, column=4, padx=10, pady=5, sticky=tk.EW)

        # Lamp indicator (Twice the original size)
        self.lamp = tk.Canvas(self.main_frame, width=100, height=100)
        self.lamp.grid(row=0, column=5, padx=20, pady=5, rowspan=3)
        self.update_lamp('red')

        # Buttons below inputs (Blue)
        self.zero_pos_button = ttk.Button(self.main_frame, text="Zero Pos", command=self.zero_pos, style="Blue.TButton")
        self.zero_pos_button.grid(row=1, column=3, padx=5, pady=5, sticky=tk.EW)

        self.zero_force_button = ttk.Button(self.main_frame, text="Zero Forza", command=self.zero_force, style="Blue.TButton")
        self.zero_force_button.grid(row=2, column=3, padx=5, pady=5, sticky=tk.EW)

        self.save_button = ttk.Button(self.main_frame, text="Salva File", command=self.save_file, style="Blue.TButton")
        self.save_button.grid(row=1, column=4, padx=5, pady=5, sticky=tk.EW)

        # Matplotlib plot
        self.fig, self.ax = plt.subplots()
        self.ax.set_title("Forza-Spostamento")
        self.ax.set_xlabel("Spostamento [mm]")
        self.ax.set_ylabel("Forza [Kg]")
        self.canvas = FigureCanvasTkAgg(self.fig, master=root)
        self.canvas.get_tk_widget().grid(row=7, column=0, columnspan=5, rowspan=3, sticky=tk.NSEW)

        root.grid_rowconfigure(7, weight=1)
        root.grid_columnconfigure(0, weight=1)

        # Bind resizing events to dynamically adjust the plot size
        root.bind("<Configure>", self.resize_plot)

        # Closing event binding to stop serial communication
        root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Define button styles
        self.style = ttk.Style()
        self.style.configure("Blue.TButton", font=("Arial", int(12 * 1.3)), foreground="black", background="blue")
        self.style.configure("Green.TButton", font=("Arial", int(12 * 1.3)), foreground="black", background="green")
        self.style.configure("Red.TButton", font=("Arial", int(12 * 1.3)), foreground="black", background="red")
        
    def  create_numeric_entry(self, label_text, row, scale_factor, variable, editable=True):
        label = ttk.Label(self.main_frame, text=label_text, font=("Arial", int(13 * scale_factor)))
        label.grid(row=row, column=0, padx=5, pady=5, sticky=tk.E)
        entry = ttk.Entry(self.main_frame, textvariable=variable, font=("Arial", int(16 * scale_factor)), state="readonly" if not editable else "normal")
        entry.grid(row=row, column=1, padx=5, pady=5, sticky=tk.W)

    def get_com_ports(self):
        """Return a list of available COM ports."""
        ports = serial.tools.list_ports.comports()
        return [port.device for port in ports]

    def connect_serial(self, event=None):
        """Connect to the selected COM port."""
        com_port = self.com_port_var.get()
        if com_port:
            try:
                # Inizializza la connessione seriale
                self.serial_obj = serial.Serial(com_port, 9600, timeout=1)
                self.conn_status_label.config(text="CONNECT_SERIAL: Connessione seriale stabilita.")
                self.connection_attempt_start_time = None  # Reset the timer
                # Attempt to clear buffers immediately after opening the port
                self.serial_obj.reset_input_buffer()
                self.serial_obj.reset_output_buffer()

                # Avvia la lettura dati
                self.start_reading_data()
                
            except serial.SerialException:
                # Errore di connessione
                self.conn_status_label.config(text="CONNECT_SERIAL --> ERRORE DI CONNESSIONE: Ricerca connessione...")
                self.serial_obj = None
                self.connection_attempt_start_time = datetime.now()  # Start the timer
                self.schedule_reconnect()

        else:
            # Nessuna porta selezionata
            self.conn_status_label.config(text="CONNECT_SERIAL: Nessuna COM port trovata")
            self.schedule_reconnect()

    def schedule_reconnect(self):
        """Programma un tentativo di riconnessione."""
        root.after(300, self.check_connection_status)

    def check_connection_status(self):
        """Check if the connection is still valid."""
        if self.serial_obj and self.serial_obj.is_open:
            # Connessione stabilita
            self.conn_status_label.config(text="CHECK_CONNECTION_STATUS: Connessione seriale stabilita.")
            self.connection_attempt_start_time = None

        else:
            # Connessione non valida, tenta la riconnessione
            if not self.connection_attempt_start_time:
                self.connection_attempt_start_time = datetime.now()  # Inizia il timer

            elapsed_time = (datetime.now() - self.connection_attempt_start_time).total_seconds()
            if elapsed_time >= 0.3:
                self.conn_status_label.config(text="Connessione seriale non stabilita. Nuovo tentativo...")
                self.connect_serial()  # Tentativo di riconnessione
            else:
                self.schedule_reconnect()   


    def start_reading_data(self):
        """Inizia a leggere dati dalla porta seriale."""
        if self.serial_obj and self.serial_obj.is_open:
            self.reading_data = True
            self.read_data()
        else:
            self.conn_status_label.config(text="START_READING_DATA: Connessione non valida.")
            self.reading_data = False


    def read_data(self):
        """Read data from the serial port every 100ms."""
        
        """Legge i dati dalla porta seriale solo se la connessione è valida."""
        if not self.serial_obj or not self.serial_obj.is_open or not self.reading_data:
            # Ferma la lettura se la connessione è persa o se il ciclo è stato interrotto
            self.conn_status_label.config(text="READ_DATA: Connessione non valida, ciclo interrotto.")
            return

        if self.serial_obj.in_waiting > 0:
            try:
                # self.conn_status_label.config(text="READ_DATA: siamo dentro al try...")
                # Try to read position and force data
                posi_plot = self.serial_obj.readline().decode('utf-8').strip()
                force_plot = self.serial_obj.readline().decode('utf-8').strip()

                # Parse the data and apply zero correction
                pos = round(float(posi_plot) - self.zeropos, 2)  # Arrotonda a 2 cifre
                force = round(float(force_plot) - self.zeroforce, 2)  # Arrotonda a 2 cifre
                abs_pos = round(float(posi_plot), 2)  # Arrotonda a 2 cifre
                
                # Update GUI variables
                self.pos_var.set(pos)
                self.carico_var.set(force)
                self.abs_pos_var.set(float(posi_plot))

                # Store the data in arrays for plotting
                self.arrayp.append(pos)
                self.arrayf.append(force)

                if not self.stop:
                    # Update the plot with new data
                    self.ax.clear()  # Clear the previous plot
                    self.ax.plot(self.arrayp, self.arrayf)
                    self.ax.set_xlabel("Spostamento [mm]")
                    self.ax.set_ylabel("Forza [Kg]")
                    self.ax.set_title("Forza-Spostamento")
                    self.canvas.draw()

            except ValueError:
                # Handle any parsing errors (non-numeric data)
                self.conn_status_label.config(text="READ DATA: EXCEPT ERROR. Errore durante la lettura dei dati")

        # Schedule the next data reading
        root.after(50, self.read_data)  # Read data again in 50ms


    def update_lamp(self, color):
        """Update the lamp's color."""
        self.lamp.delete("all")
        self.lamp.create_oval(5, 5, 95, 95, fill=color)

    def resize_plot(self, event):
        """Adjust the plot to the new window size."""
        if event.widget == root:
            self.canvas.get_tk_widget().config(width=event.width, height=event.height - 200)
            self.canvas.draw()

    def start_test(self):
        """Start the test and record the start time."""
        self.ax.clear()
        self.update_lamp('green')
        self.stop = False
        self.arrayp.clear()
        self.arrayf.clear()
        self.start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Orario di inizio prova

    def stop_test(self):
        """Stop the test."""
        self.update_lamp('red')
        self.stop = True

    def save_file(self):
        """Save the data to a CSV file, including the start time of the test."""
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        
        if file_path:
            # Prepare data for saving
            data = {
            "Posizione [mm]": self.arrayp,
            "Forza [Kg]": self.arrayf,
            }
            df = pd.DataFrame(data)
            
            # Create a new workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            
            # Write the headers and data
            ws['A1'] = "Data e Ora"
            # Write start time in the first cell
            ws['A2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws['C1'] = "Velocità [mm/min]"
            ws['C2'] = self.speed_var.get()  # Velocità
            ws['E1'] = "Posizione [mm]"
            ws['F1'] = "Forza [Kg]"

            # Write data to the worksheet
            for idx, (pos, force) in enumerate(zip(self.arrayp, self.arrayf), start=1):
                ws.cell(row=idx + 2, column=5, value=pos)  # Colonna E per Posizione
                ws.cell(row=idx + 2, column=6, value=force)  # Colonna F per Forza

            # Save the workbook to the specified file path (without image)
            wb.save(file_path)

            # 1. Save the plot as an image
            plot_image_path = file_path.replace(".xlsx", "_plot.png")  # Save plot as a PNG with the same base name
            self.fig.savefig(plot_image_path)  # Save the current plot to an image

            # 2. Insert the plot into the Excel file
            img = Image(plot_image_path)

            img.width = img.width // 2
            img.height = img.height // 2
            img.anchor = 'H1'  # Place the image at a specific cell (e.g., H5)
            ws.add_image(img)

            # Save the workbook with the plot image
            wb.save(file_path)

            # 4. Remove the temporary plot image file
            if os.path.exists(plot_image_path):
                os.remove(plot_image_path)  # Delete the plot image after saving to Excel

            messagebox.showinfo("Salvataggio Completato", "I dati sono stati salvati correttamente.")

    def zero_pos(self):
        self.zeropos = self.pos_var.get()

    def zero_force(self):
        self.zeroforce = self.carico_var.get()

    def on_closing(self):
        """Handle the window close event."""
        if self.serial_obj is not None and self.serial_obj.is_open:
            self.serial_obj.close()
        root.destroy()


# Create and run the main application window
root = tk.Tk()
app = App(root)
root.mainloop() 