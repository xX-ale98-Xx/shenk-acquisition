import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import ttkbootstrap as ttkb
from ttkbootstrap.constants import *
import serial
from serial.tools import list_ports
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageTk


class App:
    def __init__(self, root):
        self.root = root
        self.serial_obj = None
        self.stop = True
        self.enter = True
        self.zeroforce = 0
        self.zeropos = 0
        self.arrayp = []
        self.arrayf = []
        self.start_time = None
        self.connection_attempt_start_time = None

        self.abs_pos_var = tk.DoubleVar(value=0.0)
        self.pos_var = tk.DoubleVar(value=0.0)
        self.carico_var = tk.DoubleVar(value=0.0)
        self.speed_var = tk.DoubleVar()
        self.reportName = tk.StringVar()
        self.helpLblVar = tk.StringVar(value="Collegare USB e selezionare porta COM per avviare la lettura.")

        # Style setup — must use ttkb.Style() so ttkbootstrap's theme engine
        # properly applies custom colours (especially button backgrounds on Windows)
        self.apply_styles()
        self.style = ttkb.Style()

        # Adaptive window size
        self.screen_width = root.winfo_screenwidth()
        self.screen_height = root.winfo_screenheight()
        window_width = int(self.screen_width * 0.9)
        window_height = int(self.screen_height * 0.8)
        root.title("Shenk App")
        root.geometry(f"{window_width}x{window_height}")

        # Root grid
        columnsNum = 6
        rowNum = 7
        for i in range(columnsNum):
            root.grid_columnconfigure(i, weight=1)
        root.grid_rowconfigure(0, weight=1)
        root.grid_rowconfigure(1, weight=0)
        for i in range(2, rowNum):
            root.grid_rowconfigure(i, weight=2)

        # --- HEADER ---
        self.headerFrame = ttk.Frame(root, padding=20, style='headerFrame.TFrame')
        self.headerFrame.grid(row=0, column=0, columnspan=columnsNum, sticky="nsew")
        self.headerFrame.grid_rowconfigure(0, weight=1)
        self.headerFrame.grid_columnconfigure(0, weight=1)
        self.headerFrame.grid_columnconfigure(1, weight=4)

        # Logo
        logo_path = self._resource_path("img/logo_waya.png")
        self.original_image = Image.open(logo_path)
        orig_w, orig_h = self.original_image.size
        self.new_height = int(self.screen_height / 20)
        self.aspect_ratio = orig_h / orig_w
        self.new_width = int(self.new_height / self.aspect_ratio)
        image_resized = self.original_image.resize((self.new_width, self.new_height))
        self.logo = ImageTk.PhotoImage(image_resized)

        self.logo_label = ttk.Label(self.headerFrame, image=self.logo, anchor="center", style="headerLabel.TLabel")
        self.logo_label.grid(column=0, row=0, sticky="nsew", padx=5, pady=5)

        ttk.Label(
            self.headerFrame,
            text="WayAssauto - Shenk App",
            anchor="center",
            font=("Arial", 16, "bold"),
            style="headerLabel.TLabel"
        ).grid(column=1, row=0, sticky="nsew", padx=5, pady=5)

        root.bind("<Configure>", self._on_configure)

        # Header border
        root.grid_rowconfigure(1, minsize=3, weight=0)
        self.headerBorder = ttk.Frame(root, style="headerBorder.TFrame")
        self.headerBorder.grid(row=1, column=0, columnspan=columnsNum, sticky="nsew")

        # --- PANED WINDOW ---
        self.pw = ttk.PanedWindow(root, orient='horizontal', style="custom.TPanedwindow")
        self.pw.grid(row=2, column=0, rowspan=rowNum - 2, columnspan=columnsNum, sticky="nsew")

        # --- LEFT PANEL (controls) ---
        self.buttonsFrame = ttk.Frame(self.pw, style="MyCustomFrame.TFrame", padding=20)
        self.pw.add(self.buttonsFrame, weight=2)

        for i in range(8):
            self.buttonsFrame.grid_rowconfigure(i, weight=1)
        self.buttonsFrame.grid_columnconfigure(0, weight=1)
        self.buttonsFrame.grid_columnconfigure(1, weight=2)

        # COM port row
        ttk.Label(
            self.buttonsFrame,
            text="Seleziona Porta COM:",
            font=("Arial", 16),
            anchor='e',
            style="bodyLabel.TLabel"
        ).grid(row=0, column=0, sticky="ew", padx=5, pady=5)

        self.com_ports = self.get_com_ports()
        self.com_port_var = tk.StringVar(value=self.com_ports[0] if self.com_ports else "")
        self.com_port_dropdown = ttk.Combobox(
            self.buttonsFrame,
            textvariable=self.com_port_var,
            values=self.com_ports,
            state="readonly",
            style="custom.TCombobox"
        )
        self.com_port_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.com_port_dropdown.bind("<<ComboboxSelected>>", self.connect_serial)

        # Connection status
        self.conn_status_label = ttk.Label(
            self.buttonsFrame,
            text="Ricerca connessione...",
            font=("Arial", 12),
            style="bodyLabel.TLabel"
        )
        self.conn_status_label.grid(row=1, column=0, columnspan=2, padx=5, pady=2, sticky="w")

        # Abs Position labelframe
        lblfrmAbsPos = ttk.Labelframe(self.buttonsFrame, text='Posizione Ass. [mm]', style="lblFrm.TLabelframe")
        lblfrmAbsPos.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=5, pady=3)
        ttk.Label(lblfrmAbsPos, textvariable=self.abs_pos_var, font=("Arial", 20), anchor="center", background='#F1F1F1').pack(expand=True, fill="both")

        # Position labelframe + Zero Pos
        lblfrmPos = ttk.Labelframe(self.buttonsFrame, text='Posizione [mm]', style="lblFrm.TLabelframe")
        lblfrmPos.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=5, pady=3)
        ttk.Label(lblfrmPos, textvariable=self.pos_var, font=("Arial", 20), anchor="center", background='#F1F1F1').pack(expand=True, fill="both")
        ttk.Button(lblfrmPos, text="Zero Pos", style="zeroBtn.TButton", command=self.zero_pos).pack(side="right", padx=10, pady=5)

        # Force labelframe + Zero Forza
        lblfrmForce = ttk.Labelframe(self.buttonsFrame, text='Carico [Kg]', style="lblFrm.TLabelframe")
        lblfrmForce.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=5, pady=3)
        ttk.Label(lblfrmForce, textvariable=self.carico_var, font=("Arial", 20), anchor="center", background='#F1F1F1').pack(expand=True, fill="both")
        ttk.Button(lblfrmForce, text="Zero Forza", style="zeroBtn.TButton", command=self.zero_force).pack(side="right", padx=10, pady=5)

        # Speed
        ttk.Label(
            self.buttonsFrame,
            text="Velocità [mm/min]:",
            font=("Arial", 14),
            style="bodyLabel.TLabel"
        ).grid(row=5, column=0, sticky="e", padx=5, pady=5)
        ttk.Entry(self.buttonsFrame, textvariable=self.speed_var, font=("Arial", 14)).grid(row=5, column=1, sticky="w", padx=5, pady=5)

        # Lamp + Start + Stop
        lampFrame = ttk.Frame(self.buttonsFrame, style="MyCustomFrame.TFrame")
        lampFrame.grid(row=6, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        lampFrame.grid_columnconfigure(0, weight=1)
        lampFrame.grid_columnconfigure(1, weight=2)
        lampFrame.grid_columnconfigure(2, weight=2)
        lampFrame.grid_rowconfigure(0, weight=1)

        self.lamp = tk.Canvas(lampFrame, width=80, height=80, bg='#F1F1F1', highlightthickness=0)
        self.lamp.grid(row=0, column=0, padx=10, pady=5)
        self.update_lamp('red')

        ttk.Button(lampFrame, text="Test Start", style="startBtn.TButton", command=self.start_test).grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        ttk.Button(lampFrame, text="Test Stop", style="stopBtn.TButton", command=self.stop_test).grid(row=0, column=2, padx=5, pady=5, sticky="nsew")

        # --- RIGHT PANEL (graph) ---
        self.graphFrame = ttk.Frame(self.pw, style="MyCustomFrame.TFrame", padding=20)
        self.pw.add(self.graphFrame, weight=4)
        self.graphFrame.grid_columnconfigure(0, weight=1)
        self.graphFrame.grid_rowconfigure(0, weight=1)
        self.graphFrame.grid_rowconfigure(1, weight=4)

        # Save controls row (mirrors AMSLER buttonSaveFrame)
        self.buttonSaveFrame = ttk.Frame(self.graphFrame, style="MyCustomFrame.TFrame", padding=20)
        self.buttonSaveFrame.grid(row=0, column=0, sticky="nsew")
        self.buttonSaveFrame.grid_columnconfigure(0, weight=1)
        self.buttonSaveFrame.grid_columnconfigure(1, weight=5)
        self.buttonSaveFrame.grid_rowconfigure(0, weight=1)
        self.buttonSaveFrame.grid_rowconfigure(1, weight=1)

        self.helpLbl = ttk.Label(
            self.buttonSaveFrame,
            textvariable=self.helpLblVar,
            font=("Arial", 12),
            background="#F1F1F1",
            anchor="w",
            padding=(0, 20)
        )
        self.helpLbl.grid(row=0, column=0, columnspan=2, sticky="nsew")

        ttk.Button(
            self.buttonSaveFrame,
            text="Salva Report",
            style="reportBtn.TButton",
            command=self.save_report
        ).grid(row=1, column=0, sticky="w")

        self.labelReportNameEntry = ttk.Labelframe(
            self.buttonSaveFrame,
            text="Inserire nome report:",
            style="customLabelframe.TLabelframe"
        )
        self.labelReportNameEntry.grid(row=1, column=1, sticky="nsew")
        self.labelReportNameEntry.grid_columnconfigure(0, weight=1)

        ttk.Entry(
            self.labelReportNameEntry,
            textvariable=self.reportName,
            background="lightblue",
            justify="center",
            font=("Arial", 14)
        ).grid(row=0, column=0, padx=(0, 60), sticky="nsew")

        # Graph canvas
        self.fig, self.ax = plt.subplots()
        self.ax.set_title("Forza-Spostamento")
        self.ax.set_xlabel("Spostamento [mm]")
        self.ax.set_ylabel("Forza [Kg]")

        bg_color = ttkb.Style().lookup("MyCustomFrame.TFrame", "background")
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.graphFrame)
        canvas_widget = self.canvas.get_tk_widget()
        canvas_widget.configure(bg=bg_color)
        self.fig.patch.set_facecolor(bg_color)
        canvas_widget.grid(row=1, column=0, sticky='NSEW')

        # Close event
        root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def apply_styles(self):
        """Configure all ttk styles."""
        s = ttkb.Style()

        s.configure('headerFrame.TFrame', background='#ffe0b3')
        s.configure('headerBorder.TFrame', background='#ff9900')
        s.configure('headerLabel.TLabel', background='#ffe0b3')
        s.configure('MyCustomFrame.TFrame', background='#F1F1F1', padding=20)
        s.configure('bodyLabel.TLabel', background='#F1F1F1')
        s.configure('custom.TPanedwindow', background='#C1C1C1')

        s.configure('lblFrm.TLabelframe', background='#F1F1F1', labelmargins=0)
        s.configure('lblFrm.TLabelframe.Label', background='#F1F1F1')
        s.configure('customLabelframe.TLabelframe', padding=(0, 0, 0, 20), background='#F1F1F1', borderwidth=0)
        s.configure('customLabelframe.TLabelframe.Label', background='#F1F1F1')

        s.configure('startBtn.TButton',
                    font=('Arial', 18, 'bold'),
                    background='#4CAF50',
                    foreground='white',
                    padding=(30, 20),
                    relief='flat',
                    width=10,
                    anchor='center')
        s.map('startBtn.TButton',
              background=[('active', '#388E3C')],
              foreground=[('active', 'white')])

        s.configure('stopBtn.TButton',
                    font=('Arial', 18, 'bold'),
                    background='#F44336',
                    foreground='white',
                    padding=(30, 20),
                    relief='flat',
                    width=10,
                    anchor='center')
        s.map('stopBtn.TButton',
              background=[('active', '#D32F2F')],
              foreground=[('active', 'white')])

        s.configure('zeroBtn.TButton',
                    font=('Arial', 10, 'bold'),
                    background='#007BFF',
                    foreground='white',
                    padding=(10, 10),
                    relief='flat',
                    width=10,
                    anchor='center')
        s.map('zeroBtn.TButton',
              background=[('active', '#0056b3')],
              foreground=[('active', 'white')])

        s.configure('reportBtn.TButton',
                    font=('Arial', 12, 'bold'),
                    background='#007BFF',
                    foreground='white',
                    padding=(30, 10),
                    relief='flat',
                    width=12,
                    anchor='center')
        s.map('reportBtn.TButton',
              background=[('active', '#0056b3')],
              foreground=[('active', 'white')])

        s.configure('custom.TCombobox', font=("Arial", 12), width=10)

        # Base TFrame background — must be set so all frames inherit #F1F1F1
        # and the headerFrame override (#ffe0b3) is applied on top
        s.configure('TFrame', background='#F1F1F1')
        s.configure('headerFrame.TFrame', background='#ffe0b3')

    def _resource_path(self, relative_path):
        """Resolve path for both development and PyInstaller bundle."""
        try:
            base_path = sys._MEIPASS
        except AttributeError:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def _on_configure(self, event):
        """Handle window resize: update logo and plot."""
        if event.widget == self.root:
            self._resize_logo()

    def _resize_logo(self):
        """Dynamically resize logo based on current window height."""
        new_h = int(self.root.winfo_height() / 20)
        if new_h < 1:
            return
        new_w = int(new_h / self.aspect_ratio)
        resized = self.original_image.resize((new_w, new_h), Image.LANCZOS)
        self.logo_tk = ImageTk.PhotoImage(resized)
        self.logo_label.config(image=self.logo_tk)
        self.logo_label.image = self.logo_tk

    # -------------------------------------------------------------------------
    # Serial / data logic — identical to v4
    # -------------------------------------------------------------------------

    def get_com_ports(self):
        """Return a list of available COM ports."""
        ports = serial.tools.list_ports.comports()
        return [port.device for port in ports]

    def connect_serial(self, event=None):
        """Connect to the selected COM port."""
        com_port = self.com_port_var.get()
        if com_port:
            try:
                self.serial_obj = serial.Serial(com_port, 9600, timeout=1)
                self.conn_status_label.config(text="CONNECT_SERIAL: Connessione seriale stabilita.")
                self.connection_attempt_start_time = None
                self.serial_obj.reset_input_buffer()
                self.serial_obj.reset_output_buffer()
                self.start_reading_data()
            except serial.SerialException:
                self.conn_status_label.config(text="CONNECT_SERIAL --> ERRORE DI CONNESSIONE: Ricerca connessione...")
                self.serial_obj = None
                self.connection_attempt_start_time = datetime.now()
                self.schedule_reconnect()
        else:
            self.conn_status_label.config(text="CONNECT_SERIAL: Nessuna COM port trovata")
            self.schedule_reconnect()

    def schedule_reconnect(self):
        """Schedule a reconnection attempt."""
        self.root.after(300, self.check_connection_status)

    def check_connection_status(self):
        """Check if the connection is still valid."""
        if self.serial_obj and self.serial_obj.is_open:
            self.conn_status_label.config(text="CHECK_CONNECTION_STATUS: Connessione seriale stabilita.")
            self.connection_attempt_start_time = None
        else:
            if not self.connection_attempt_start_time:
                self.connection_attempt_start_time = datetime.now()
            elapsed_time = (datetime.now() - self.connection_attempt_start_time).total_seconds()
            if elapsed_time >= 0.3:
                self.conn_status_label.config(text="Connessione seriale non stabilita. Nuovo tentativo...")
                self.connect_serial()
            else:
                self.schedule_reconnect()

    def start_reading_data(self):
        """Start reading data from the serial port."""
        if self.serial_obj and self.serial_obj.is_open:
            self.reading_data = True
            self.read_data()
        else:
            self.conn_status_label.config(text="START_READING_DATA: Connessione non valida.")
            self.reading_data = False

    def read_data(self):
        """Read data from the serial port every 50ms."""
        if not self.serial_obj or not self.serial_obj.is_open or not self.reading_data:
            self.conn_status_label.config(text="READ_DATA: Connessione non valida, ciclo interrotto.")
            return

        if self.serial_obj.in_waiting > 0:
            try:
                posi_plot = self.serial_obj.readline().decode('utf-8').strip()
                force_plot = self.serial_obj.readline().decode('utf-8').strip()

                pos = round(float(posi_plot) - self.zeropos, 2)
                force = round(float(force_plot) - self.zeroforce, 2)

                self.pos_var.set(pos)
                self.carico_var.set(force)
                self.abs_pos_var.set(float(posi_plot))

                self.arrayp.append(pos)
                self.arrayf.append(force)

                if not self.stop:
                    self.ax.clear()
                    self.ax.plot(self.arrayp, self.arrayf)
                    self.ax.set_xlabel("Spostamento [mm]")
                    self.ax.set_ylabel("Forza [Kg]")
                    self.ax.set_title("Forza-Spostamento")
                    self.canvas.draw()

            except ValueError:
                self.conn_status_label.config(text="READ DATA: EXCEPT ERROR. Errore durante la lettura dei dati")

        self.root.after(50, self.read_data)

    def update_lamp(self, color):
        """Update the lamp indicator colour."""
        self.lamp.delete("all")
        self.lamp.create_oval(5, 5, 75, 75, fill=color)

    def start_test(self):
        """Start the test."""
        self.ax.clear()
        self.update_lamp('green')
        self.stop = False
        self.arrayp.clear()
        self.arrayf.clear()
        self.start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def stop_test(self):
        """Stop the test."""
        self.update_lamp('red')
        self.stop = True

    def save_report(self):
        """Save data to Excel — same format as v4 (PNG plot embedded)."""
        report_name = self.reportName.get().strip()
        if not report_name:
            messagebox.showerror("Errore", "Inserire un nome per il report.")
            return
        if not self.stop:
            messagebox.showerror("Errore", "Fermare la prova prima di salvare il report.")
            return
        if not self.arrayp or not self.arrayf:
            messagebox.showerror("Errore", "Nessun dato da salvare.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Salva report come...",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=report_name
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active

        ws['A1'] = "Data e Ora"
        ws['A2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['C1'] = "Velocità [mm/min]"
        ws['C2'] = self.speed_var.get()
        ws['E1'] = "Posizione [mm]"
        ws['F1'] = "Forza [Kg]"

        for idx, (pos, force) in enumerate(zip(self.arrayp, self.arrayf), start=1):
            ws.cell(row=idx + 2, column=5, value=pos)
            ws.cell(row=idx + 2, column=6, value=force)

        wb.save(file_path)

        plot_image_path = file_path.replace(".xlsx", "_plot.png")
        self.fig.savefig(plot_image_path)

        img = XLImage(plot_image_path)
        img.width = img.width // 2
        img.height = img.height // 2
        img.anchor = 'H1'
        ws.add_image(img)

        wb.save(file_path)

        if os.path.exists(plot_image_path):
            os.remove(plot_image_path)

        self.show_user_message(f"Report salvato: {os.path.basename(file_path)}", timeout=4000)
        messagebox.showinfo("Salvataggio Completato", "I dati sono stati salvati correttamente.")

    def show_user_message(self, message, timeout=3000):
        """Show a temporary message in the help label."""
        self.helpLblVar.set(message)
        if hasattr(self, "_msg_timer") and self._msg_timer:
            self.root.after_cancel(self._msg_timer)
        self._msg_timer = self.root.after(timeout, self._restore_status_message)

    def _restore_status_message(self):
        """Restore the help label to the current serial connection status."""
        if self.serial_obj and self.serial_obj.is_open:
            self.helpLblVar.set("Comunicazione seriale attiva: lettura in corso...")
        else:
            self.helpLblVar.set("Seriale non connessa. Collegare USB e selezionare porta COM.")

    def zero_pos(self):
        self.zeropos = self.abs_pos_var.get()

    def zero_force(self):
        self.zeroforce = self.carico_var.get() + self.zeroforce

    def on_closing(self):
        """Handle window close."""
        if self.serial_obj is not None and self.serial_obj.is_open:
            self.serial_obj.close()
        self.root.destroy()


if __name__ == "__main__":
    root = ttkb.Window(themename="simplex")
    app = App(root)
    root.mainloop()
